from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook
import requests
from openpyxl import Workbook


file='prova.xlsx'


book = Workbook()
wb = load_workbook(file)
ws = wb.active

tim = 45

a=0
#citta = ['Germany','russia','united kingdom', 'france','italy','spain', 'poland','Netherlands','belgium', 'sweden', 'belarus', 'austria','switzerland','denmark','finland','norway','ireland','montenegro','luxembourg','malta','iceland','andorra','monaco','liechtenstein','san marino','gibraltar' ]
citta=['France']
#citta = ['Canada','italy', 'poland','belgium','spain','Netherlands', 'sweden', 'austria','switzerland','denmark','finland', 'ireland','norway','montenegro','luxembourg','malta','iceland','andorra','monaco','liechtenstein','san marino','gibraltar', 'chile', 'japan', 'israel', 'new zeland']
print(len(citta))
kkk=0
#citta = ['Netherlands']
#jobs = ['financial advisor','business analyst','consulting']
#jobs = ['Conservatory', 'Start Up Incubator', 'Music', 'venture capital', 'social house entertainment','influencer marketing']
#jobs = ['social house entertainment']
jobs = ['influencer marketing','Music', 'venture capital','social house entertainment']

checkNoDoppioni = []

contatore = 0
#citta = ['poland','belarus','austria']
# city = nation
for job in jobs:
    for state in citta:
        d = len(citta)
        lista = []
        checkNoDoppioni = []
        if kkk == 1:
            locations_btn = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, '//span//button[contains(@aria-label,"Locations filter")]'))).click()
            search = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[6]/div[3]/div[2]/section/div/nav/div/ul/li[3]/div/div/div/div[1]/div/form/fieldset/div[1]/div/div/input")))
            search.send_keys(state)
            time.sleep(1.5)
            search.send_keys(Keys.ARROW_DOWN)
            time.sleep(0.5)
            search.send_keys(Keys.ENTER)
            time.sleep(0.5)
            locations_btn = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[6]/div[3]/div[2]/section/div/nav/div/ul/li[3]/div/div/div/div[1]/div/form/fieldset/div[2]/button[2]'))).click()
            time.sleep(2)
            locations_btn = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, '//span//button[contains(@aria-label,"Locations filter")]')))
            color = locations_btn.value_of_css_property("backgroundColor")
            color = str(color)
            if color == 'rgba(5, 118, 66, 1)':
                driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                time.sleep(0.9)
                sp = BeautifulSoup(driver.page_source, 'html.parser')
                links = sp.find_all('li', class_='reusable-search__result-container')
                time.sleep(0.9)
                if len(links) > 0:  # IF THERE IS AT LEAST ONE RESULT
                    for x in links:
                        gg = x.find('a', class_='app-aware-link').get('href')
                        if gg not in checkNoDoppioni:
                            checkNoDoppioni.append(gg)
                            lista.append(gg)  # THIS IS A LIST OF RESULTS LINKS
                    try:

                        # WHILE THE 'NEXT'BUTTON IS ENABLE:
                        while WebDriverWait(driver, 8).until(EC.visibility_of_element_located((By.XPATH,
                                                                                               "//button[contains(@aria-label, 'Next')]//*[contains(., 'Next')]/.."))).is_enabled():
                            driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                            time.sleep(0.9)
                            WebDriverWait(driver, tim).until(EC.visibility_of_element_located(
                                (By.XPATH,
                                 "//button[contains(@aria-label, 'Next')]//*[contains(., 'Next')]/.."))).click()
                            time.sleep(2)
                            sp = BeautifulSoup(driver.page_source, 'html.parser')
                            links = sp.find_all('li', class_='reusable-search__result-container')
                            for x in links:
                                gg = x.find('a', class_='app-aware-link').get('href')
                                if gg not in lista: lista.append(gg)
                                time.sleep(0.5)
                                driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                                time.sleep(0.9)
                        print('terminato per questa città')

                    except:
                        print('solo una pagina')
                        print(len(lista))
                        print(lista)
                else:
                    print('no results for this state/city')
                    print(lista)
                a +=1
                for x in lista:
                    ws.append([job,state, state, x])
                wb.save(file)

                # CLEARING THE RESULTS
                WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[6]/div[3]/div[2]/section/div/nav/div/div/button"))).click()
                print(str(a) + '/' + str(d))
            else:
                print('questa città non è selezionabile')
                driver.quit()
                kkk=0
        # so if kkk != 0  since this is not the first city we are looking for
        else:
            kkk =1
            d = len(citta)
            #  LOGIN
            driver = webdriver.Chrome(ChromeDriverManager().install())
            driver.get("https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin")
            time.sleep(3)
            main_window = driver.current_window_handle
            username = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='session_key']")))
            password = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "//input[@name='session_password']")))
            username.send_keys('')
            password.send_keys('')
            time.sleep(0.5)
            submit = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "//button[@type='submit']"))).click()


            # LOOKING FOR THE FIRST JOB
            search = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "//*[@id='global-nav-typeahead']/input")))
            search.send_keys(job)
            search.send_keys(Keys.ENTER)


            # GETTING COMPANIES FILTER
            search = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[6]/div[3]/div[2]/section/div/nav/div/ul/li[2]/button'))).click()

            # GETTING COUNTRIES FILTER
            locations_btn = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, '//span//button[contains(@aria-label,"Locations filter")]'))).click()
            search = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[6]/div[3]/div[2]/section/div/nav/div/ul/li[3]/div/div/div/div[1]/div/form/fieldset/div[1]/div/div/input")))
            search.send_keys(state)
            time.sleep(0.9)
            search.send_keys(Keys.ARROW_DOWN)
            time.sleep(0.8)
            search.send_keys(Keys.ENTER)
            time.sleep(1.5)

            #SENDING REQUEST TO GET DATA FOR JOB IN A PARTICULAR COUNTRY
            locations_btn = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[6]/div[3]/div[2]/section/div/nav/div/ul/li[3]/div/div/div/div[1]/div/form/fieldset/div[2]/button[2]'))).click()
            time.sleep(2)

            #CHECKING IF THERE IS AT LEAST A RESULT FOR THIS RESEARCH
            locations_btn = WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, '//span//button[contains(@aria-label,"Locations filter")]')))
            color = locations_btn.value_of_css_property("backgroundColor")
            color = str(color)
            print('risultati trovati')

                # IF THE BUTTON COLOR IS GREEN, IT MEANS THAT THERE IS AT LEAST A VALUABLE RESULT
            if color == 'rgba(5, 118, 66, 1)':
                driver.execute_script("window.scrollTo(0,document.body.scrollHeight)") #SCROLLING TO THE BOTTOM OF THE PAGE
                time.sleep(0.9)
                sp = BeautifulSoup(driver.page_source, 'html.parser')
                links = sp.find_all('li', class_='reusable-search__result-container')
                time.sleep(0.9)
                if len(links) > 0: # IF THERE IS AT LEAST ONE RESULT
                    for x in links:
                        gg = x.find('a', class_='app-aware-link').get('href')
                        if gg not in checkNoDoppioni:
                            checkNoDoppioni.append(gg)
                            lista.append(gg) # THIS IS A LIST OF RESULTS LINKS
                    try:

                            # WHILE THE 'NEXT'BUTTON IS ENABLE:
                        while WebDriverWait(driver, 8).until(EC.visibility_of_element_located((By.XPATH, "//button[contains(@aria-label, 'Next')]//*[contains(., 'Next')]/.."))).is_enabled():
                            driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                            time.sleep(0.9)
                            WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "//button[contains(@aria-label, 'Next')]//*[contains(., 'Next')]/.."))).click()
                            time.sleep(1.3)
                            sp = BeautifulSoup(driver.page_source, 'html.parser')
                            links = sp.find_all('li', class_='reusable-search__result-container')
                            for x in links:
                                gg = x.find('a', class_='app-aware-link').get('href')
                                if gg not in lista: lista.append(gg)
                                time.sleep(0.5)
                                driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
                                time.sleep(0.9)
                        print('terminato per questa città')

                    except:
                        print('solo una pagina')
                        print(len(lista))
                        print(lista)
                else: # NON  CI SONO RISULTATI PER QUESTA CITTA'
                    print('no results for this state/city')
                    print(lista)
                a += 1
                # lista is a list with all the links of linkedin's prifiles
                for x in lista:
                    ws.append([job,state, state, x])
                wb.save(file)
                print(str(a) + '/' + str(d))


                #CLEARING THE FILTERS
                WebDriverWait(driver, tim).until(EC.visibility_of_element_located((By.XPATH, "/html/body/div[6]/div[3]/div[2]/section/div/nav/div/div/button"))).click()

                # IF THE BUTTON COLOR IS NOT GREEN, IT MEANS THAT WE NEED TO GO ON WITH A NEW COUNTRY ---> NO RESULT FOR THAT NATION
            else:
                print('città non identificabile')
                driver.quit()
                kkk = 0
        contatore = 1
    kkk =0
    driver.quit()

wb.save(file)

