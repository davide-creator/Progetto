from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook

# Ora di inizio: 20:30 Roma

# what do I need?
# - site's link
# - position
# - title


# skip deve essere pari alla somma di tutte le iterazioni eseguite
skip = 74563




file = r'prova.xlsx'
filePath = r'overtuneNoDopp.xlsx'

db = pd.read_excel(filePath)
numero_link = len(db)



db = db.loc[skip:numero_link]


book = Workbook()
wb = load_workbook(file)
ws = wb.active

gg = 0
xxx = 1


for n in range(numero_link):
    link = 555555555555555555555555555555555555555555
    job = 555555555555555555555555555555555555555555
    state = 555555555555555555555555555555555555555555
    city = 555555555555555555555555555555555555555555
    title = 555555555555555555555555555555555555555555
    siteLink = 555555555555555555555555555555555555555555
    position = 555555555555555555555555555555555555555555
    x = db.loc[n+skip]
    link = x['url']
    job = x['job']
    state = x['state']
    city = x['city']
    print(gg,'/',numero_link)
    print(gg+skip, '/', numero_link)
    if xxx == 0:  # non è la prima ricerca
        try:
            driver.get(link)
            WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
                (By.XPATH, "//button[contains(@aria-label, 'Follow')]//*[contains(., 'Follow')]/..")))
            sp = BeautifulSoup(driver.page_source, 'html.parser')
            # site link:
            try:
                siteLink = sp.find('a', class_= 'ember-view org-top-card-primary-actions__action').get('href')
                print('siteLink: ',siteLink)
            except:
                print('no website available')
                siteLink = ''
            # position
            try:
                position = sp.find('div', class_= 'org-top-card-summary-info-list t-14 t-black--light')
                pos = []
                try:
                    inizio = position.find('div', class_='org-top-card-summary-info-list__info-item').get_text()
                    inizio = inizio.strip()
                    pos.append(inizio)
                except: print('non ci sono div di posizione')
                try:
                    inline = position.find('div', 'inline-block')
                    for x in inline.find_all('div'):
                        testo = x.get_text()
                        testo = testo.strip()
                        pos.append(testo)
                except: print('non ci sono din inline')
                if len(pos) > 0: position = ' '.join(pos)
                else: position = ''
                print('position: ',position)
            except:
                print('no position available')
                position = ''
            # title
            try:#                             ember-view t-24 t-black t-bold

                title = sp.find('h1', class_='ember-view t-24 t-black t-bold full-width').get_text()
                title = title.strip()
                print('title: ', title)
            except:
                print('no title available')
                title = ''
        except: print('link linkedin non più valido')
        gg+=1
        if siteLink != '' and siteLink != 555555555555555555555555555555555555555555:
            ws.append([job, state, city, title, link, siteLink, position])
            wb.save(file)
        else: print('contatto eliminato perchè senza link')
        xxx = 0
        print(job, state, city, title, link, siteLink, position)

    if xxx == 1:
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get("https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin")
        time.sleep(3)
        username = WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@name='session_key']")))
        password = WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@name='session_password']")))
        username.send_keys('davide.balestra3@studio.unibo.it')
        password.send_keys('Elefante12')
        submit = WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, "//button[@type='submit']"))).click()
        time.sleep(2.5)

        try:
            driver.get(link)
            try:
                WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
                    (By.XPATH, "//button[contains(@aria-label, 'Follow')]//*[contains(., 'Follow')]/..")))
                sp = BeautifulSoup(driver.page_source, 'html.parser')
                # site link:
                try:
                    siteLink = sp.find('a', class_='ember-view org-top-card-primary-actions__action').get('href')
                    print('siteLink: ', siteLink)
                except:
                    print('no website available')
                    siteLink = ''
                # position
                try:
                    position = sp.find('div', class_='org-top-card-summary-info-list t-14 t-black--light')
                    pos = []
                    try:
                        inizio = position.find('div', class_='org-top-card-summary-info-list__info-item').get_text()
                        inizio = inizio.strip()
                        pos.append(inizio)
                    except: print('non ci sono div di posizione')
                    try:
                        inline = position.find('div', 'inline-block')
                        for x in inline.find_all('div'):
                            testo = x.get_text()
                            testo = testo.strip()
                            pos.append(testo)
                    except: print('non ci sono din inline')
                    if len(pos) > 0:
                        position = ' '.join(pos)
                    else:
                        position = ''
                    print('position: ', position)
                except:
                    print('no position available')
                    position = ''
                # title
                try:
                    title = sp.find('h1', class_='ember-view t-24 t-black t-bold full-width').get_text()
                    title = title.strip()
                    print('title: ', title)
                except:
                    print('no title available')
                    title = ''
            except: print('cacciati dal sito')
        except: print('link linkedin non più valido')
        gg += 1
        if siteLink != '' and siteLink != 555555555555555555555555555555555555555555:
            ws.append([job, state, city, title, link, siteLink, position])
            wb.save(file)
        else:
            print('contatto eliminato perchè senza link')
        xxx = 0
        print(job, state, city, title, link, siteLink, position)
driver.quit()