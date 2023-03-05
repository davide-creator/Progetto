import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd
import pdfkit
import os
import time
from random import randrange


data = 'May 19, 2022'
skip = 0
testo = 'testoEmail.txt'
# TODO: change the file with emails
file_con_email = 'inviareEmail.xlsx'
SMTP_SERVER = "smtp.office365.com"
SMTP_PORT = 587
sender_email ='davide.balestra@outlook.com'
password = 'Elefante12'
email_inviate = 'email_inviate.xlsx'
email_inviate_errori = 'Email_inviate_con_errori.xlsx'

book = Workbook()
inviate = load_workbook(email_inviate)
sheet_inviate = inviate.active

errori = load_workbook(email_inviate_errori)
sheet_errori = errori.active

db = pd.read_excel(file_con_email)
contatti = len(db)
print(contatti)

# TODO: change contatti with the amount of mails you wish to send
for x in range(skip,contatti): # range(start, stop)
    #tempo = randrange(95,800)
    time.sleep(randrange(25,68))
    try:
        try: os.remove('Balestra Davide Cover Letter.pdf')
        except: print('invio prima email...')
        
        panda_dato = db.loc[x]
        receiver_email = panda_dato[2]
        
        name = panda_dato[5]
        try: name = name.replace(',','')
        except: a=1

        if type(receiver_email) == type(name) and type(name) != 'nan':     
            msg = MIMEMultipart("alternative")
            #TODO: change the subject of the email:
            msg["Subject"] = 'Application for a summer internship'
            msg["From"] = sender_email
            msg["To"] = receiver_email
            with open(testo, 'r') as f:
                message = f.read()            
                # Replace the target string
                message = message.replace('NAME$$', name)
            
            part = MIMEText(message, "plain")
            msg.attach(part)

            # Add Attachment

            # first file
            f = open('Balestra Davide Cover Letter.htm', 'r')
            f = f.read()            
            f = f.replace('EMAIL$$', 'davide.balestra1@yahoo.com')
            f = f.replace('DATE$$', data)     
            f = f.replace('NAME$$', name)
            pdfkit.from_string(f, 'Balestra Davide Cover Letter.pdf')
            

            
            with open('Balestra Davide Cover Letter.pdf', "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())        
            encoders.encode_base64(part)
           

            # second file
            part.add_header(
                "Content-Disposition",
                "attachment", filename= 'Balestra Davide Cover Letter.pdf'
            )
            msg.attach(part)


            with open('Balestra Davide Curriculum.pdf'  , "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
            
            encoders.encode_base64(part)


            # Set mail headers
            part.add_header(
                "Content-Disposition",
                "attachment", filename= 'Balestra Davide Curriculum.pdf'  
            )
            msg.attach(part)

            # Create secure SMTP connection and send email
            context = ssl.create_default_context()
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()          
                server.login(sender_email, password)
                server.sendmail(sender_email, receiver_email, msg.as_string())

                print('success...  ' + str(x) + ' ' + receiver_email)
                sheet_inviate.append([panda_dato[0],panda_dato[1],panda_dato[2],panda_dato[3],panda_dato[4],panda_dato[5],panda_dato[6],panda_dato[7]])
                inviate.save(email_inviate)
            
        else: 
            print('indirizzio email non attendibile... ' + receiver_email )
            sheet_errori.append([panda_dato[0],panda_dato[1],panda_dato[2],panda_dato[3],panda_dato[4],panda_dato[5],panda_dato[6],panda_dato[7]])
            errori.save(email_inviate_errori)
        
    except:
        sheet_errori.append([panda_dato[0],panda_dato[1],panda_dato[2],panda_dato[3],panda_dato[4],panda_dato[5],panda_dato[6],panda_dato[7]])
        errori.save(email_inviate_errori)
        print('la mail fallita è:...' + panda_dato[2])